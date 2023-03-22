from openpyxl import load_workbook

from src.params import Params
from train_random_forest import train_random_forest
from train_generation import train_generation
from train_embedding import train_embedding


# ===== Pre training =====

# Data
data_small = './data/training_data_small.pkl'
data_small_aug = './data/training_data_small_augmented.pkl'

data_medium = './data/training_data_generation.pkl'
data_medium_aug = './data/training_data_generation_augmented.pkl'

data_large = './data/training_data.pkl'
data_large_aug = './data/training_data_augmented.pkl'

# Models
model_path = 'result_model/cbow_network.pt'
model_path_aug = 'result_model/cbow_network_augmented.pt'

# Runs
small_run_pre = {'nb_of_runs': 5, 'column': 'D'}
medium_run_pre = {'nb_of_runs': 3, 'column': 'E'}
large_run_pre = {'nb_of_runs': 1, 'column': 'F'}

# Must do embeddings first.
pretrain_runs = [
    # Embddings 
    {'program': train_embedding, 'run_instance': [large_run_pre], 'row': 67, 'embed': '', 'shuffle': '', 'slide_range': '','generation_data': '', 'embed_data': data_large, 
        'model_path': model_path},
    {'program': train_embedding, 'run_instance': [large_run_pre], 'row': 68, 'embed': '', 'shuffle': '', 'slide_range': '','generation_data': '', 'embed_data': data_large_aug, 
        'model_path': model_path_aug},

    # LSTM Embed
    {'program': train_generation, 'run_instance': [small_run_pre], 'row': 41, 'embed': True, 'shuffle': '', 'slide_range': 1,'generation_data': data_small, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_generation, 'run_instance': [small_run_pre], 'row': 42, 'embed': True, 'shuffle': '', 'slide_range': 1,'generation_data': data_small_aug, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_generation, 'run_instance': [small_run_pre], 'row': 44, 'embed': True, 'shuffle': '', 'slide_range': 5,'generation_data': data_small, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_generation, 'run_instance': [small_run_pre], 'row': 45, 'embed': True, 'shuffle': '', 'slide_range': 5,'generation_data': data_small_aug, 'embed_data': '', 
        'model_path': model_path_aug},

    # LSTM Vanilla
    {'program': train_generation, 'run_instance': [small_run_pre], 'row': 48, 'embed': False, 'shuffle': '', 'slide_range': 1,'generation_data': data_small, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_generation, 'run_instance': [small_run_pre], 'row': 49, 'embed': False, 'shuffle': '', 'slide_range': 1,'generation_data': data_small_aug, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_generation, 'run_instance': [small_run_pre], 'row': 51, 'embed': False, 'shuffle': '', 'slide_range': 5,'generation_data': data_small, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_generation, 'run_instance': [small_run_pre], 'row': 52, 'embed': False, 'shuffle': '', 'slide_range': 5,'generation_data': data_small_aug, 'embed_data': '', 
        'model_path': model_path_aug},

    # Hidden Forest Embed
    {'program': train_random_forest, 'run_instance': [small_run_pre], 'row': 55, 'embed': True, 'shuffle': True, 'slide_range': 1,'generation_data': data_small, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': [small_run_pre], 'row': 56, 'embed': True, 'shuffle': True, 'slide_range': 1,'generation_data': data_small_aug, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': [small_run_pre], 'row': 58, 'embed': True, 'shuffle': True, 'slide_range': 5,'generation_data': data_small, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': [small_run_pre], 'row': 59, 'embed': True, 'shuffle': True, 'slide_range': 5,'generation_data': data_small_aug, 'embed_data': '', 
        'model_path': model_path_aug},

    # Hidden Forest Vanilla
    {'program': train_random_forest, 'run_instance': [small_run_pre], 'row': 62, 'embed': False, 'shuffle': True, 'slide_range': 1,'generation_data': data_small, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': [small_run_pre], 'row': 63, 'embed': False, 'shuffle': True, 'slide_range': 1,'generation_data': data_small_aug, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': [small_run_pre], 'row': 65, 'embed': False, 'shuffle': True, 'slide_range': 5,'generation_data': data_small, 'embed_data': '', 
        'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': [small_run_pre], 'row': 66, 'embed': False, 'shuffle': True, 'slide_range': 5,'generation_data': data_small_aug, 'embed_data': '', 
        'model_path': model_path_aug},
    
]

# Uncomment for pre training.
# for run in pretrain_runs:
#     for run_instance in run['run_instance']:
#         results = []

#         for run_index in range(run_instance['nb_of_runs']):
#             params = Params(embed_data=run['embed'],
#                             shuffle_data_random_forest=run['shuffle'],
#                             window_slide_range=run['slide_range'],
#                             generation_training_data_path=run['generation_data'],
#                             embedding_training_data_path=run['embed_data'],
#                             embedding_model_path=run['model_path']
#                             )
            
#             result = run['program'](params)
#             results.append(result)

#         # Format result. Cannot put '=' at start because excel think it's hacking.
#         results = [str(r) for r in results]
#         result = ' + '.join(results)
#         result = result.replace('.', ',')
#         result = f'({result}) / {len(results)} * 100'

#         # Write to excel
#         workbook = load_workbook(filename = 'results.xlsx')
#         sheet = workbook.active

#         cell = f"{run_instance['column']}{str(run['row'])}"
#         sheet[cell] = result
#         workbook.save('results.xlsx')
#         print(f'saved cell {cell}')


# ==== Results training ======
small_run = {'nb_of_runs': 5, 'column': 'E', 'data': {'normal': data_small, 'augmented': data_small_aug}}
medium_run = {'nb_of_runs': 3, 'column': 'F', 'data': {'normal': data_medium, 'augmented': data_medium_aug}}
large_run = {'nb_of_runs': 1, 'column': 'G', 'data': {'normal': data_large, 'augmented': data_large_aug}}

all_runs = [small_run, medium_run, large_run]

runs = [
    ### LSTM

    # LSTM Embed
    {'program': train_generation, 'run_instance': all_runs, 'row': 4, 'embed': True, 'shuffle': False, 'data': 'normal', 'model_path': model_path},
    {'program': train_generation, 'run_instance': all_runs, 'row': 5, 'embed': True, 'shuffle': False, 'data': 'normal', 'model_path': model_path_aug},
    {'program': train_generation, 'run_instance': all_runs, 'row': 6, 'embed': True, 'shuffle': False, 'data': 'augmented', 'model_path': model_path},
    {'program': train_generation, 'run_instance': all_runs, 'row': 7, 'embed': True, 'shuffle': False, 'data': 'augmented', 'model_path': model_path_aug},

    # LSTM Vanilla
    {'program': train_generation, 'run_instance': all_runs, 'row': 9, 'embed': False, 'shuffle': False, 'data': 'normal', 'model_path': model_path},
    {'program': train_generation, 'run_instance': all_runs, 'row': 10, 'embed': False, 'shuffle': False, 'data': 'normal', 'model_path': model_path_aug},
    {'program': train_generation, 'run_instance': all_runs, 'row': 11, 'embed': False, 'shuffle': False, 'data': 'augmented', 'model_path': model_path},
    {'program': train_generation, 'run_instance': all_runs, 'row': 12, 'embed': False, 'shuffle': False, 'data': 'augmented', 'model_path': model_path_aug},

    ### Random Forest

    # Random Forest Embed - No Shuffle
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 15, 'embed': True, 'shuffle': False, 'data': 'normal', 'model_path': model_path},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 16, 'embed': True, 'shuffle': False, 'data': 'normal', 'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 17, 'embed': True, 'shuffle': False, 'data': 'augmented', 'model_path': model_path},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 18, 'embed': True, 'shuffle': False, 'data': 'augmented', 'model_path': model_path_aug},

    # Random Forest Embed - Shuffle
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 20, 'embed': True, 'shuffle': True, 'data': 'normal', 'model_path': model_path},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 21, 'embed': True, 'shuffle': True, 'data': 'normal', 'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 22, 'embed': True, 'shuffle': True, 'data': 'augmented', 'model_path': model_path},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 23, 'embed': True, 'shuffle': True, 'data': 'augmented', 'model_path': model_path_aug},

    # Random Forest Vanilla - No Shuffle
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 26, 'embed': False, 'shuffle': False, 'data': 'normal', 'model_path': model_path},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 27, 'embed': False, 'shuffle': False, 'data': 'normal', 'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 28, 'embed': False, 'shuffle': False, 'data': 'augmented', 'model_path': model_path},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 29, 'embed': False, 'shuffle': False, 'data': 'augmented', 'model_path': model_path_aug},

    # Random Forest Vanilla - Shuffle
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 31, 'embed': False, 'shuffle': True, 'data': 'normal', 'model_path': model_path},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 32, 'embed': False, 'shuffle': True, 'data': 'normal', 'model_path': model_path_aug},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 33, 'embed': False, 'shuffle': True, 'data': 'augmented', 'model_path': model_path},
    {'program': train_random_forest, 'run_instance': all_runs, 'row': 34, 'embed': False, 'shuffle': True, 'data': 'augmented', 'model_path': model_path_aug},

]
        
for run in runs:
    for run_instance in run['run_instance']:
        results = []

        for run_index in range(run_instance['nb_of_runs']):
            data = run_instance['data'][run['data']]

            params = Params(embed_data=run['embed'],
                            shuffle_data_random_forest=run['shuffle'],
                            window_slide_range=1,   # Alway use full window slide
                            generation_training_data_path=data,
                            embedding_model_path=run['model_path']
                            )
            
            result = run['program'](params)
            results.append(result)

        # Format result. Cannot put '=' at start because excel think it's hacking.
        results = [str(r) for r in results]
        result = ' + '.join(results)
        result = result.replace('.', ',')
        result = f'({result}) / {len(results)} * 100'

        # Write to excel
        workbook = load_workbook(filename = 'results.xlsx')
        sheet = workbook.active

        cell = f"{run_instance['column']}{str(run['row'])}"
        sheet[cell] = result
        workbook.save('results.xlsx')
        print(f'saved cell {cell}')